import openpyxl

# Abre o arquivo modelo
arquivo_modelo = "modelo-eventos.xlsx"
wb = openpyxl.load_workbook(arquivo_modelo)
ws = wb["Planilha1"]

# Preenche com dados de teste
ws["A2"] = "20 crianças"
ws["B2"] = "10/04/2026 - 14h00"
ws["C2"] = "Rua Exemplo, 123"
ws["A3"] = "Maria Silva"  # Cliente (corrigido)
ws["A1"] = "Tema Safari"  # Tema (provavelmente aqui, baseado na mesclagem A1:B1)
ws["A4"] = "Oficina de Slime"  # Pacote
ws["B6"] = "Cliente pediu atenção com crianças menores."
ws["B7"] = "R$ 850,00"

# Salva novo arquivo
novo_arquivo = "ficha-evento-preenchida.xlsx"
wb.save(novo_arquivo)

print("Arquivo preenchido com sucesso!")