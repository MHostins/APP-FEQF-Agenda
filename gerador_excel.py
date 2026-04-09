import openpyxl


def gerar_excel(evento, nome_arquivo="ficha_evento.xlsx"):
    arquivo_modelo = "modelo-eventos.xlsx"
    wb = openpyxl.load_workbook(arquivo_modelo)
    ws = wb["Planilha1"]

    ws["A2"] = evento.get("quantidade", "")
    ws["B2"] = evento.get("data_horario", "")
    ws["C2"] = evento.get("endereco", "")
    ws["A1"] = evento.get("tema", "")
    ws["A3"] = evento.get("cliente", "")
    ws["A4"] = evento.get("pacote", "")
    ws["B6"] = evento.get("observacao", "")
    ws["B7"] = evento.get("total", "")

    wb.save(nome_arquivo)
    return nome_arquivo